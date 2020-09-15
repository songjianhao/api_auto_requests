import os, sys

sys.path.append(os.path.dirname(os.path.dirname(__file__)))

import shutil
from openpyxl import load_workbook
from openpyxl.styles import Font, Alignment
from openpyxl.styles.colors import RED, GREEN, DARKYELLOW

name = "SongJianhao"

class write_excel():
    def __init__(self, file_name):
        self.file_name = file_name
        if not os.path.exists(self.file_name):
            '''文件不存在则将测试用例复制到测试结果文件'''
            shutil.copyfile()

        self.workbook = load_workbook(self.file_name)
        self.worksheet = self.workbook.active

    def write_data(self, row_n, reBody, result):
        '''
        将测试结果写入报告文件
        传入当前行数 row_n
        传入当前返回值reBody
        传入当前测试结果result
        '''
        font_green = Font(name='宋体', color=GREEN, bold=True)
        font_red = Font(name='宋体', color=RED, bold=True)
        font_default = Font(name='宋体')
        font_sign = Font(name='宋体', color=DARKYELLOW, bold=True)
        align = Alignment(horizontal='center', vertical='center')
        # 获取当前行数
        K_n = "k" + str(row_n)
        L_n = "L" + str(row_n)
        M_n = "M" + str(row_n)

        self.worksheet.cell(row_n, 11, reBody)
        self.worksheet[K_n].font = font_default

        if result == "PASS":
            self.worksheet.cell(row_n, 12, result)
            self.worksheet[L_n].font = font_green
        if result == "FAIL":
            self.worksheet.cell(row_n, 12, result)
            self.worksheet[L_n].font = font_red
        self.worksheet[L_n].alignment = align

        self.worksheet.cell(row_n, 13, name)
        self.worksheet[M_n].font = font_sign
        self.worksheet[M_n].alignment = align

        self.workbook.save(self.file_name)



